import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

excel_file = "issues_report.xlsx"

class XEngine:

    def __init__(self):
        # Get the path to the .openai directory in the user's home directory
        headers = ["Issue", "Level", "Created", "Description"]
        df = pd.DataFrame(columns=headers)
        df.to_excel(excel_file, index=False)
        wb = load_workbook(excel_file)
        ws = wb.active
        aheader = ws['A1']
        bheader = ws['B1']
        cheader = ws['C1']
        dheader = ws['D1']
        aheader.alignment = Alignment(horizontal='left')
        bheader.alignment = Alignment(horizontal='left')
        cheader.alignment = Alignment(horizontal='left')
        dheader.alignment = Alignment(horizontal='left')
        ws.column_dimensions['A'].width = 12  # Issue column
        ws.column_dimensions['B'].width = 12  # Level column
        ws.column_dimensions['C'].width = 12  # Created column
        ws.column_dimensions['D'].width = 250  # Description column
        wb.save(excel_file)
        print(f"Excel file '{excel_file}' created successfully with three data sets, titles in blue, and duplicated data.")

    def initialize(self, title, data):
        wb = load_workbook(excel_file)
        ws = wb.active
        num_rows = ws.max_row
        ws.merge_cells(f'A{num_rows + 1}:C{num_rows + 1}')  # Merging columns A, B, C
        merged_cell = ws[f'A{num_rows + 1}']
        merged_cell.value = title
        merged_cell.alignment = Alignment(horizontal='left')
        merged_cell.font = Font(bold=True, color="0000FF")  # Blue font color in hex
        for i, row in enumerate(data, start=num_rows + 2):
            ws.append(row)
        wb.save(excel_file)

if __name__ == "__main__":
    xengine = XEngine()
    data = [
        ["ISSUE-001", "Critical", "01/01/2024", "System outage in region A"],
        ["ISSUE-002", "High", "01/02/2024", "Database connection issue"],
        ["ISSUE-003", "Medium", "01/03/2024", "Minor UI bug in dashboard"],
        ["ISSUE-004", "Low", "01/04/2024", "Typo in help documentation"]
    ]

    xengine.initialize('AKUNA-MATADA', data)
    xengine.initialize('BOODIE-HOLE', data)